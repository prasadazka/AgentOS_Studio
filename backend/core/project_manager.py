"""Project CRUD, file processing, sessions, and messages for AgentOS Studio."""

import asyncio
import json
import os
import shutil
import threading
from pathlib import Path
from typing import Any, Optional

from db.database import get_db, generate_id, now_iso
from core.memory_pool import get_project_memory, evict_project_memory

PROJECTS_DIR = Path(os.getenv("AGENTOS_PROJECTS_DIR", os.path.expanduser("~/.agent_os/projects")))


# ---------------------------------------------------------------------------
# Projects
# ---------------------------------------------------------------------------

def create_project(name: str, description: str, agent_name: str) -> dict[str, Any]:
    db = get_db()
    pid = generate_id()
    ts = now_iso()
    db.execute(
        "INSERT INTO projects (id, name, description, agent_name, created_at, updated_at) VALUES (?,?,?,?,?,?)",
        (pid, name, description, agent_name, ts, ts),
    )
    db.commit()
    return {"id": pid, "name": name, "description": description, "agent_name": agent_name, "created_at": ts, "updated_at": ts}


def list_projects() -> list[dict[str, Any]]:
    db = get_db()
    rows = db.execute(
        """SELECT p.*,
                  (SELECT COUNT(*) FROM project_files WHERE project_id = p.id) AS file_count,
                  (SELECT COUNT(*) FROM chat_sessions WHERE project_id = p.id) AS session_count
           FROM projects p ORDER BY p.updated_at DESC"""
    ).fetchall()
    return [dict(r) for r in rows]


def get_project(pid: str) -> Optional[dict[str, Any]]:
    db = get_db()
    row = db.execute(
        """SELECT p.*,
                  (SELECT COUNT(*) FROM project_files WHERE project_id = p.id) AS file_count,
                  (SELECT COUNT(*) FROM chat_sessions WHERE project_id = p.id) AS session_count
           FROM projects p WHERE p.id = ?""",
        (pid,),
    ).fetchone()
    return dict(row) if row else None


def delete_project(pid: str) -> bool:
    db = get_db()
    row = db.execute("SELECT id FROM projects WHERE id = ?", (pid,)).fetchone()
    if not row:
        return False
    db.execute("DELETE FROM projects WHERE id = ?", (pid,))
    db.commit()
    # Clean up files on disk
    project_dir = PROJECTS_DIR / pid
    if project_dir.exists():
        shutil.rmtree(project_dir, ignore_errors=True)
    # Evict memory
    evict_project_memory(pid)
    return True


# ---------------------------------------------------------------------------
# Files
# ---------------------------------------------------------------------------

async def add_file_streamed(project_id: str, filename: str, upload_file) -> dict[str, Any]:
    """Stream-upload a file to disk without loading it all into RAM, then process in background."""
    db = get_db()
    fid = generate_id()
    ts = now_iso()

    ext = Path(filename).suffix.lower()
    file_type = ext.lstrip(".")

    files_dir = PROJECTS_DIR / project_id / "files"
    files_dir.mkdir(parents=True, exist_ok=True)
    filepath = files_dir / f"{fid}_{filename}"

    # Stream to disk in 1 MB chunks
    file_size = 0
    CHUNK = 1024 * 1024  # 1 MB
    with open(filepath, "wb") as f:
        while True:
            chunk = await upload_file.read(CHUNK)
            if not chunk:
                break
            f.write(chunk)
            file_size += len(chunk)

    # Insert DB record immediately so UI can show the file
    db.execute(
        "INSERT INTO project_files (id, project_id, filename, filepath, file_type, file_size, status, uploaded_at) VALUES (?,?,?,?,?,?,?,?)",
        (fid, project_id, filename, str(filepath), file_type, file_size, "processing", ts),
    )
    db.commit()

    # Process (chunk + embed) in background thread so the HTTP response returns immediately
    def _process():
        _embed_file(fid, project_id, filename, filepath, ext)
        db2 = get_db()
        db2.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (now_iso(), project_id))
        db2.commit()

    threading.Thread(target=_process, daemon=True).start()

    return dict(db.execute("SELECT * FROM project_files WHERE id = ?", (fid,)).fetchone())


def _embed_file(fid: str, project_id: str, filename: str, filepath: Path, ext: str):
    """Extract text, chunk, and embed a file. Called from background thread."""
    db = get_db()
    try:
        text = _extract_text(filepath, ext)
        if text.strip():
            memory = get_project_memory(project_id)
            chunks = _chunk_text(text, chunk_size=1000, overlap=200)
            for i, chunk in enumerate(chunks):
                memory.store_knowledge(
                    content=chunk,
                    metadata={"filename": filename, "file_id": fid, "chunk_index": i},
                )
        db.execute("UPDATE project_files SET status = 'ready' WHERE id = ?", (fid,))
    except Exception as e:
        db.execute("UPDATE project_files SET status = 'error' WHERE id = ?", (fid,))
        print(f"[project_manager] File processing error for {filename}: {e}")
    db.commit()


def add_file(project_id: str, filename: str, content_bytes: bytes) -> dict[str, Any]:
    """Save uploaded file, extract text, chunk, and embed into project memory."""
    db = get_db()
    fid = generate_id()
    ts = now_iso()

    # Determine file type
    ext = Path(filename).suffix.lower()
    file_type = ext.lstrip(".")

    # Save to disk
    files_dir = PROJECTS_DIR / project_id / "files"
    files_dir.mkdir(parents=True, exist_ok=True)
    filepath = files_dir / f"{fid}_{filename}"
    filepath.write_bytes(content_bytes)

    # Insert DB record
    db.execute(
        "INSERT INTO project_files (id, project_id, filename, filepath, file_type, file_size, status, uploaded_at) VALUES (?,?,?,?,?,?,?,?)",
        (fid, project_id, filename, str(filepath), file_type, len(content_bytes), "processing", ts),
    )
    db.commit()

    _embed_file(fid, project_id, filename, filepath, ext)

    # Update project timestamp
    db.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (now_iso(), project_id))
    db.commit()

    return dict(db.execute("SELECT * FROM project_files WHERE id = ?", (fid,)).fetchone())


def list_files(project_id: str) -> list[dict[str, Any]]:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM project_files WHERE project_id = ? ORDER BY uploaded_at DESC", (project_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def delete_file(file_id: str) -> bool:
    db = get_db()
    row = db.execute("SELECT filepath, project_id FROM project_files WHERE id = ?", (file_id,)).fetchone()
    if not row:
        return False
    # Delete from disk
    fp = Path(row["filepath"])
    if fp.exists():
        fp.unlink()
    db.execute("DELETE FROM project_files WHERE id = ?", (file_id,))
    db.commit()
    return True


def list_all_disk_files(project_id: str) -> list[dict[str, Any]]:
    """List ALL files on disk for a project — uploaded + agent-generated.

    Returns a flat list with relative paths from the project root.
    """
    project_dir = PROJECTS_DIR / project_id / "files"
    if not project_dir.exists():
        return []

    # DB files indexed by filepath for quick lookup
    db_files = {f["filepath"]: f for f in list_files(project_id)}

    results = []
    for fp in sorted(project_dir.rglob("*")):
        if fp.is_dir():
            continue
        fp_str = str(fp)
        rel_path = str(fp.relative_to(project_dir))

        # Check if this file is in DB (uploaded)
        db_entry = db_files.get(fp_str)
        if db_entry:
            results.append({
                "id": db_entry["id"],
                "filename": db_entry["filename"],
                "filepath": fp_str,
                "relative_path": rel_path,
                "file_type": db_entry["file_type"],
                "file_size": db_entry["file_size"],
                "status": db_entry["status"],
                "source": "uploaded",
            })
        else:
            # Agent-generated file (not in DB)
            stat = fp.stat()
            results.append({
                "id": None,
                "filename": fp.name,
                "filepath": fp_str,
                "relative_path": rel_path,
                "file_type": fp.suffix.lstrip("."),
                "file_size": stat.st_size,
                "status": "ready",
                "source": "generated",
            })

    return results


def get_file_path(project_id: str, filename: str) -> Optional[Path]:
    """Resolve a filename to its actual path on disk (safe, no traversal)."""
    project_dir = PROJECTS_DIR / project_id / "files"
    if not project_dir.exists():
        return None

    # Try exact match first
    target = project_dir / filename
    if target.exists() and target.is_file():
        # Security: ensure it's within project dir
        try:
            target.resolve().relative_to(project_dir.resolve())
            return target
        except ValueError:
            return None

    # Search by filename suffix (for DB files stored as {id}_{filename})
    for fp in project_dir.rglob("*"):
        if fp.is_file() and fp.name.endswith(filename):
            try:
                fp.resolve().relative_to(project_dir.resolve())
                return fp
            except ValueError:
                continue

    return None


# ---------------------------------------------------------------------------
# Sessions
# ---------------------------------------------------------------------------

def create_session(project_id: str, title: str = "New Chat") -> dict[str, Any]:
    db = get_db()
    sid = generate_id()
    ts = now_iso()
    db.execute(
        "INSERT INTO chat_sessions (id, project_id, title, created_at, updated_at) VALUES (?,?,?,?,?)",
        (sid, project_id, title, ts, ts),
    )
    db.commit()
    return {"id": sid, "project_id": project_id, "title": title, "created_at": ts, "updated_at": ts}


def list_sessions(project_id: str) -> list[dict[str, Any]]:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM chat_sessions WHERE project_id = ? ORDER BY updated_at DESC", (project_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def rename_session(session_id: str, title: str) -> Optional[dict[str, Any]]:
    db = get_db()
    row = db.execute("SELECT id FROM chat_sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return None
    ts = now_iso()
    db.execute("UPDATE chat_sessions SET title = ?, updated_at = ? WHERE id = ?", (title, ts, session_id))
    db.commit()
    return dict(db.execute("SELECT * FROM chat_sessions WHERE id = ?", (session_id,)).fetchone())


def delete_session(session_id: str) -> bool:
    db = get_db()
    row = db.execute("SELECT id FROM chat_sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return False
    db.execute("DELETE FROM chat_sessions WHERE id = ?", (session_id,))
    db.commit()
    return True


def delete_all_sessions(project_id: str) -> int:
    """Delete all chat sessions for a project. Returns count of deleted sessions."""
    db = get_db()
    rows = db.execute("SELECT id FROM chat_sessions WHERE project_id = ?", (project_id,)).fetchall()
    count = len(rows)
    if count > 0:
        db.execute("DELETE FROM chat_sessions WHERE project_id = ?", (project_id,))
        db.commit()
    return count


# ---------------------------------------------------------------------------
# Messages
# ---------------------------------------------------------------------------

def add_message(
    session_id: str,
    role: str,
    content: str,
    tool_calls: Optional[list] = None,
) -> dict[str, Any]:
    db = get_db()
    mid = generate_id()
    ts = now_iso()
    # Safety: Gemini models may return content as a list of parts — ensure string
    if isinstance(content, list):
        content = "\n".join(
            p.get("text", str(p)) if isinstance(p, dict) else str(p)
            for p in content
        )
    tc_json = json.dumps(tool_calls) if tool_calls else None
    db.execute(
        "INSERT INTO chat_messages (id, session_id, role, content, tool_calls_json, created_at) VALUES (?,?,?,?,?,?)",
        (mid, session_id, role, content, tc_json, ts),
    )
    # Update session timestamp
    db.execute("UPDATE chat_sessions SET updated_at = ? WHERE id = ?", (ts, session_id))
    db.commit()
    return {"id": mid, "session_id": session_id, "role": role, "content": content, "tool_calls_json": tc_json, "created_at": ts}


def get_messages(session_id: str) -> list[dict[str, Any]]:
    db = get_db()
    rows = db.execute(
        "SELECT * FROM chat_messages WHERE session_id = ? ORDER BY created_at ASC", (session_id,)
    ).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_text(filepath: Path, ext: str) -> str:
    """Extract text content from a file."""
    if ext in (".csv", ".txt", ".md", ".json", ".log", ".py", ".js", ".ts", ".yaml", ".yml"):
        return filepath.read_text(encoding="utf-8", errors="replace")

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(filepath), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    parts.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(parts)
        except ImportError:
            return f"[Cannot read .xlsx — openpyxl not installed]"

    if ext == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(filepath))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            return f"[Cannot read .pdf — PyPDF2 not installed]"

    return filepath.read_text(encoding="utf-8", errors="replace")


def _chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> list[str]:
    """Split text into overlapping chunks."""
    if len(text) <= chunk_size:
        return [text]
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap
    return chunks
